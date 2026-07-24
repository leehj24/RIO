from flask import Flask, request, jsonify, send_from_directory
from pathlib import Path
import json
import datetime
import threading
from openpyxl import Workbook, load_workbook
from config import Settings
from toss_client import TossClient, TossAPIError
from auth import TossAuthError
from strategy.events import load_events
from strategy.symbol_registry import load_symbols, load_watchlist_raw

app = Flask(__name__, static_folder="dashboard")


@app.errorhandler(TossAuthError)
def handle_toss_auth_error(error):
    return jsonify({"error": "toss_auth_error", "message": str(error)}), 502


@app.errorhandler(TossAPIError)
def handle_toss_api_error(error):
    return jsonify({"error": "toss_api_error", "message": str(error)}), 502


def client():
    s = Settings()
    return TossClient(
        base_url=s.toss_base_url,
        access_token=None,
        default_account_seq=s.toss_account_seq,
        dry_run=s.dry_run,
        auto_account=s.toss_auto_account,
    )


@app.get("/")
def index():
    return send_from_directory("dashboard", "index.html")


@app.get("/app.js")
def app_js():
    return send_from_directory("dashboard", "app.js")


@app.get("/api/token")
def token():
    return jsonify(client().issue_token())


@app.get("/api/accounts")
def accounts():
    return jsonify(client().accounts())


@app.get("/api/buying-power")
def buying_power():
    return jsonify(client().buying_power(account_seq=request.args.get("account")))


@app.get("/api/positions")
def positions():
    c = client()
    data = c.positions(account_seq=request.args.get("account"))
    if request.args.get("profit") == "1":
        result = data.get("result", data)
        if isinstance(result, list):
            data["_with_estimated_profit"] = [c.estimate_profit_fields(x) for x in result]
    return jsonify(data)


@app.get("/api/stocks")
def stocks():
    symbols = [s.strip() for s in request.args.get("symbols", "").split(",") if s.strip()]
    return jsonify(client().stocks(symbols))


@app.get("/api/prices")
def prices():
    symbols = [s.strip() for s in request.args.get("symbols", "").split(",") if s.strip()]
    return jsonify(client().prices(symbols))


@app.get("/api/candles")
def candles():
    return jsonify(client().candles(
        request.args.get("symbol", ""),
        interval=request.args.get("interval", "1d"),
        count=int(request.args.get("count", "120")),
    ))


@app.get("/api/stock/analyze")
def stock_analyze():
    symbol = request.args.get("symbol", "").strip()
    force_refresh = request.args.get("refresh", "0") == "1"
    if not symbol:
        return jsonify({"error": "Symbol is required"}), 400

    from strategy.symbol_registry import describe_symbol
    from strategy.llm_providers import GoogleGeminiGroundingClient
    from config import Settings

    s = Settings()
    sym_info = describe_symbol(symbol)
    name = sym_info.get("name", symbol)
    market = sym_info.get("market", "KR")

    # Check cache first
    cache_dir = Path("data_cache/ai_analysis")
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_file = cache_dir / f"{symbol}.json"

    if not force_refresh and cache_file.exists():
        try:
            cached_data = json.loads(cache_file.read_text(encoding="utf-8"))
            return jsonify({"symbol": symbol, "name": name, "result": cached_data, "cached": True})
        except Exception as e:
            print(f"Error reading cache for {symbol}: {e}")

    # If no cache and we are not forcing a refresh/generation, return empty result to let the frontend show an opt-in button
    if not cache_file.exists() and not force_refresh and request.args.get("generate", "0") != "1":
        return jsonify({"symbol": symbol, "name": name, "result": None, "cached": False})

    # If no cache or force refresh, call LLM
    api_key = (s.google_api_key or s.gemini_api_key or "").strip()
    if not api_key:
        import os
        api_key = (os.getenv("GOOGLE_API_KEY") or os.getenv("GEMINI_API_KEY") or "").strip()

    if not api_key:
        return jsonify({
            "error": "no_api_key",
            "message": "구글 제미나이 API 키가 설정되지 않았습니다. .env 파일을 확인해 주세요."
        }), 400

    client_llm = GoogleGeminiGroundingClient(
        api_key=api_key,
        model=s.gemini_model or "gemini-2.5-flash",
        enable_grounding=s.enable_google_grounding
    )

    prompt = f"""
너는 투자 비전공자도 쉽게 이해할 수 있도록 친근하고 명쾌하게 설명해주는 주식 리서치 분석가다.
대상 종목: {name} (티커/코드: {symbol}, 시장: {market})

최신 인터넷 검색 기능(Google Search Grounding)을 적극적으로 활용하여 아래의 항목들을 정성적으로 분석해 다정하게 설명해 줘. 너무 딱딱하거나 기술적인 용어, 무거운 숫자 위주의 분석은 피하고 비전공자도 "아, 그렇구나!" 할 수 있게 친숙한 말투(존댓말, '~해요'체 혹은 격식체 존칭)로 답해 줘.

반드시 아래 JSON 스키마 형식으로만 반환해 줘. (JSON 외의 다른 텍스트는 응답에 포함하지 말 것)

JSON Schema:
{{
  "company_intro": "이 회사가 어떤 일을 하고 어떤 제품/서비스를 만드는지 비전공자용 친숙한 설명 (2-3문장)",
  "macro_analysis": "앞으로의 금리 전망, 글로벌 환율 흐름, 미국/글로벌 정책 등의 경제 환경이 이 종목에 구체적으로 어떤 영향을 주는지에 대한 쉬운 설명 (예: '앞으로 미국의 금리가 내리면 이 회사는...', '앞으로 미국이 무역 규제를 하면 이 회사는...' 등, 3-4문장)",
  "vision": "결과적으로 이 회사가 어떠한 비전과 성장 가능성을 가지고 있고, 앞으로 어떻게 나아갈 회사인지에 대한 긍정적이고 객관적인 비전 전망 (2-3문장)",
  "risk_level": "매우안전 / 안전 / 위험 / 매우위험 중 딱 하나만 선택",
  "risk_reason": "위 위험등급(risk_level)을 선택한 구체적이고 설득력 있는 이유 (비전공자 관점, 2-3문장)"
}}
"""

    try:
        result = client_llm.generate_json(prompt)
        
        # Save to cache
        try:
            cache_file.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as e:
            print(f"Error writing cache for {symbol}: {e}")
            
        return jsonify({"symbol": symbol, "name": name, "result": result, "cached": False})
    except Exception as e:
        return jsonify({"error": "llm_error", "message": str(e)}), 500


@app.post("/api/order")
def order():
    body = request.json or {}
    return jsonify(client().create_order(**body))


@app.post("/api/modify")
def modify():
    body = request.json or {}
    return jsonify(client().modify_order(**body))


@app.post("/api/cancel")
def cancel():
    body = request.json or {}
    return jsonify(client().cancel_order(**body))


# Excel Log lock and helper
excel_lock = threading.Lock()

def log_api_to_excel(path, method, query_params, request_body, status_code, response_text):
    now = datetime.datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")
    
    log_dir = Path("log")
    log_dir.mkdir(parents=True, exist_ok=True)
    excel_path = log_dir / "api_logs.xlsx"
    
    # Format params & body
    query_str = json.dumps(query_params, ensure_ascii=False) if query_params else ""
    body_str = ""
    if request_body:
        if isinstance(request_body, bytes):
            try:
                body_str = request_body.decode('utf-8', errors='ignore')
            except Exception:
                body_str = str(request_body)
        else:
            body_str = json.dumps(request_body, ensure_ascii=False)
            
    # Truncate large responses to fit excel cell limit (32767 chars)
    if response_text and len(response_text) > 10000:
        response_text = response_text[:10000] + "... (truncated)"

    row_data = [time_str, path, method, query_str, body_str, status_code, response_text]
    
    with excel_lock:
        try:
            if excel_path.exists():
                wb = load_workbook(excel_path)
            else:
                wb = Workbook()
                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])
            
            if date_str in wb.sheetnames:
                ws = wb[date_str]
            else:
                ws = wb.create_sheet(title=date_str)
                ws.append(["Time", "Path", "Method", "Query Params", "Request Body", "Status Code", "Response Content"])
            
            ws.append(row_data)
            wb.save(excel_path)
            wb.close()
        except Exception as e:
            print(f"Error logging API to excel: {e}", flush=True)


@app.after_request
def log_after_request(response):
    if request.path.startswith("/api/"):
        try:
            req_body = None
            if request.is_json:
                req_body = request.get_json(silent=True)
            elif request.data:
                req_body = request.data
                
            resp_text = None
            if response.is_json:
                resp_text = response.get_data(as_text=True)
            else:
                resp_text = f"Non-JSON response (mimetype: {response.mimetype})"
                
            log_api_to_excel(
                path=request.path,
                method=request.method,
                query_params=dict(request.args),
                request_body=req_body,
                status_code=response.status_code,
                response_text=resp_text
            )
        except Exception as e:
            print(f"Error in request logger hook: {e}", flush=True)
            
    return response


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=True)


@app.get("/api/bot/status")
def bot_status():
    s = Settings()
    state_path = Path(s.state_path)
    state = {}
    if state_path.exists():
        try:
            state = json.loads(state_path.read_text(encoding="utf-8"))
        except Exception as e:
            state = {"error": str(e)}

    log_path = Path(s.log_path)
    last_log = None
    if log_path.exists():
        try:
            lines = log_path.read_text(encoding="utf-8").splitlines()
            if lines:
                last_log = json.loads(lines[-1])
        except Exception as e:
            last_log = {"error": str(e)}

    return jsonify({
        "dry_run": s.dry_run,
        "loop_seconds": s.loop_seconds,
        "state": state,
        "last_log": last_log,
    })


@app.get("/api/bot/trades")
def bot_trades():
    s = Settings()
    n = int(request.args.get("n", "50"))
    p = Path(s.log_path)
    if not p.exists():
        return jsonify({"result": []})

    rows = []
    for line in p.read_text(encoding="utf-8").splitlines()[-n:]:
        try:
            rows.append(json.loads(line))
        except Exception:
            rows.append({"raw": line})
    return jsonify({"result": rows})


@app.get("/api/bot/events")
def bot_events():
    events = load_events()
    return jsonify({
        "result": [
            {
                "event_id": e.event_id,
                "theme": e.theme,
                "question": e.question,
                "mapped_symbols": e.mapped_symbols,
                "b": e.b,
            }
            for e in events
        ]
    })


@app.get("/api/bot/symbols")
def bot_symbols():
    # The dashboard only exposes symbols that can enter the live Toss path.
    return jsonify({"result": load_symbols()})


@app.get("/api/bot/execution-symbols")
def bot_execution_symbols():
    """Return only enabled KR/US symbols that can enter the Toss order path."""
    return jsonify({"result": load_symbols()})


@app.get("/api/bot/watchlist-raw")
def bot_watchlist_raw():
    return jsonify({"result": load_watchlist_raw()})


@app.get("/api/bot/llm-calls")
def bot_llm_calls():
    s = Settings()
    n = int(request.args.get("n", "50"))
    p = Path(s.llm_log_path)
    if not p.exists():
        return jsonify({"result": []})

    rows = []
    for line in p.read_text(encoding="utf-8").splitlines()[-n:]:
        try:
            rows.append(json.loads(line))
        except Exception:
            rows.append({"raw": line})
    return jsonify({"result": rows})


@app.get("/api/bot/api-usage")
def bot_api_usage():
    s = Settings()
    p = Path(s.api_usage_path)
    if not p.exists():
        return jsonify({
            "date": None,
            "providers": {
                "google": {"limit": s.google_daily_call_limit, "used": 0, "remaining": s.google_daily_call_limit},
                "nvidia": {"limit": s.nvidia_daily_call_limit, "used": 0, "remaining": s.nvidia_daily_call_limit},
                "gemini_api2": {"limit": s.gemini_api2_daily_call_limit, "used": 0, "remaining": s.gemini_api2_daily_call_limit},
            }
        })

    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except Exception as e:
        return jsonify({"error": str(e)})

    providers = data.get("providers", {})
    out = {"date": data.get("date"), "providers": {}}
    for provider, limit in {
        "google": s.google_daily_call_limit,
        "nvidia": s.nvidia_daily_call_limit,
        "gemini_api2": s.gemini_api2_daily_call_limit,
        "naver_search": getattr(s, "naver_search_daily_call_limit", 300),
    }.items():
        used = int(providers.get(provider, {}).get("count", 0))
        out["providers"][provider] = {
            "limit": limit,
            "used": used,
            "remaining": max(0, limit - used),
            "calls": providers.get(provider, {}).get("calls", []),
        }
    return jsonify(out)


@app.get("/api/bot/data-requirements")
def bot_data_requirements():
    import csv
    p = Path("data/actual_data_requirements.csv")
    if not p.exists():
        return jsonify({"result": []})
    with p.open("r", encoding="utf-8-sig", newline="") as f:
        return jsonify({"result": list(csv.DictReader(f))})


# ------------------------------------------------------------
# 자동투자 시작/정지 제어 + 한글 요약
# ------------------------------------------------------------

BOT_CONTROL_PATH = Path("data/bot_control.json")


def read_bot_control() -> dict:
    if BOT_CONTROL_PATH.exists():
        try:
            return json.loads(BOT_CONTROL_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"run": True, "updated_at": None}


@app.get("/api/bot/control")
def bot_control_get():
    return jsonify(read_bot_control())


@app.post("/api/bot/control")
def bot_control_set():
    body = request.get_json(silent=True) or {}
    control = {
        "run": bool(body.get("run", True)),
        "updated_at": datetime.datetime.now().isoformat(timespec="seconds"),
    }
    BOT_CONTROL_PATH.parent.mkdir(parents=True, exist_ok=True)
    BOT_CONTROL_PATH.write_text(json.dumps(control, ensure_ascii=False, indent=2), encoding="utf-8")
    return jsonify(control)


@app.get("/api/bot/summary")
def bot_summary():
    """대시보드용 한글 요약: 체결 내역·판단 통계·수익률·보유종목."""
    s = Settings()

    state = {}
    state_path = Path(s.state_path)
    if state_path.exists():
        try:
            state = json.loads(state_path.read_text(encoding="utf-8"))
        except Exception:
            state = {}

    executed = []
    decision_counts = {}
    today = datetime.date.today().isoformat()
    today_buys = today_sells = 0

    p = Path(s.log_path)
    if p.exists():
        # 최근 2000줄만 스캔 (파일이 수백 MB일 수 있음)
        try:
            lines = p.read_text(encoding="utf-8").splitlines()[-2000:]
        except Exception:
            lines = []
        for line in lines:
            try:
                rec = json.loads(line)
            except Exception:
                continue
            action = rec.get("action")
            if action:
                decision_counts[action] = decision_counts.get(action, 0) + 1
            execution = rec.get("execution") or {}
            if execution.get("submitted") or execution.get("executed"):
                ts = str(rec.get("ts", ""))
                item = {
                    "ts": ts,
                    "symbol": rec.get("symbol"),
                    "name": (rec.get("symbol_info") or {}).get("name"),
                    "action": action,
                    "price": rec.get("price"),
                    "order": rec.get("order_spec") or execution.get("order") or {},
                    "reason": rec.get("action_reason"),
                }
                executed.append(item)
                if ts[:10] == today:
                    if "BUY" in str(action).upper():
                        today_buys += 1
                    elif "SELL" in str(action).upper():
                        today_sells += 1

    return jsonify({
        "control": read_bot_control(),
        "dry_run": s.dry_run,
        "today": today,
        "today_buys": today_buys,
        "today_sells": today_sells,
        "executed_recent": executed[-30:][::-1],
        "decision_counts": decision_counts,
        "performance": {
            "daily_pnl_pct": state.get("daily_pnl_pct", 0.0),
            "total_drawdown_pct": state.get("total_drawdown_pct", 0.0),
            "portfolio_equity_krw": state.get("portfolio_equity_krw"),
            "portfolio_peak_equity_krw": state.get("portfolio_peak_equity_krw"),
            "api_error_count": state.get("api_error_count", 0),
        },
        "positions_state": state.get("positions", {}),
    })
